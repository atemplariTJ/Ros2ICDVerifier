import os
from typing import List
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from core.models import TopicInfo
from config.settings import ValidationStatus


# ── Badge text helpers (동일 로직을 table_model과 공유) ────────────────────

def _reception_text(t: TopicInfo) -> str:
    return "수신됨" if t.received else "미수신"

def _qos_text(t: TopicInfo) -> str:
    pub = t.actual_pub_qos or "-"
    sub = t.actual_sub_qos or "-"
    if pub in ("-", "Unknown"):
        return "-"
    if pub.lower() != t.target_qos.lower():
        return "불일치"
    if pub == "BestEffort" and sub == "Reliable":
        return "비호환"
    if sub not in ("-", "Unknown") and sub.lower() != t.target_qos.lower():
        return "불일치"
    return "정상"

def _hz_text(t: TopicInfo) -> str:
    if t.target_hz == 0:
        return "비주기·정상" if t.received else "미수신"
    if t.status == ValidationStatus.HZ_MISMATCH:
        return "주기미달"
    if t.actual_hz and t.actual_hz > 0:
        return "정상"
    return "대기중"

def _summary_text(t: TopicInfo) -> str:
    mapping = {
        ValidationStatus.NORMAL:       "정상",
        ValidationStatus.PENDING:      "대기중",
        ValidationStatus.NOT_RECEIVED: "미수신",
        ValidationStatus.QOS_MISMATCH: "QoS불일치",
        ValidationStatus.HZ_MISMATCH:  "주기미달",
        ValidationStatus.ERROR:        "오류",
    }
    return mapping.get(t.status, t.status.value)


# ── Cell fill colors (UI 배지 색상과 동일) ────────────────────────────────

_FILL = {
    "green":  PatternFill("solid", fgColor="DCFCE7"),
    "orange": PatternFill("solid", fgColor="FFEDD5"),
    "yellow": PatternFill("solid", fgColor="FEF9C3"),
    "pink":   PatternFill("solid", fgColor="FCE7F3"),
    "blue":   PatternFill("solid", fgColor="EFF6FF"),
    "gray":   PatternFill("solid", fgColor="F3F4F6"),
    "red":    PatternFill("solid", fgColor="FEE2E2"),
    "white":  PatternFill("solid", fgColor="FFFFFF"),
    "header": PatternFill("solid", fgColor="1E3A5F"),
}

_FONT = {
    "green":  Font(color="166534", bold=True),
    "orange": Font(color="9A3412", bold=True),
    "yellow": Font(color="854D0E", bold=True),
    "pink":   Font(color="9D174D", bold=True),
    "blue":   Font(color="1D4ED8", bold=True),
    "gray":   Font(color="4B5563"),
    "red":    Font(color="991B1B", bold=True),
    "header": Font(color="FFFFFF", bold=True),
}

def _badge_style(text: str):
    """배지 텍스트로부터 (fill, font) 반환."""
    if text in ("정상", "수신됨", "비주기·정상"):
        return _FILL["green"], _FONT["green"]
    if text in ("불일치", "비호환", "QoS불일치"):
        return _FILL["orange"], _FONT["orange"]
    if text == "주기미달":
        return _FILL["yellow"], _FONT["yellow"]
    if text in ("미수신",):
        return _FILL["pink"], _FONT["pink"]
    if text == "비주기·정상":
        return _FILL["blue"], _FONT["blue"]
    if text in ("대기중", "-"):
        return _FILL["gray"], _FONT["gray"]
    if text == "오류":
        return _FILL["red"], _FONT["red"]
    return _FILL["white"], Font()


# ── 컬럼 정의 ──────────────────────────────────────────────────────────────

# (헤더명, 값추출함수, badge여부)
_COLUMNS = [
    ("토픽명",       lambda t: t.name,                                          False),
    ("타입",         lambda t: t.topic_type,                                     False),
    ("송신 노드",    lambda t: ", ".join(t.connected_publishers) or "-",         False),
    ("수신 노드",    lambda t: ", ".join(t.connected_subscribers) or "-",        False),
    ("RobotID 송신", lambda t: t.header_src_id or "-",                           False),
    ("RobotID 수신", lambda t: ", ".join(t.header_dst_ids) or "-",               False),
    ("목표 QoS",     lambda t: t.target_qos,                                     False),
    ("송신 QoS",     lambda t: t.actual_pub_qos or "-",                          False),
    ("수신 QoS",     lambda t: t.actual_sub_qos or "-",                          False),
    ("목표 Hz",      lambda t: "비주기" if t.target_hz == 0 else str(t.target_hz), False),
    ("실제 Hz",      lambda t: f"{t.actual_hz:.2f}" if t.actual_hz else "-",     False),
    # ── 검증 상태 (UI 배지와 동일) ──
    ("수신",         _reception_text,  True),
    ("QoS",          _qos_text,        True),
    ("주기",         _hz_text,         True),
    ("종합",         _summary_text,    True),
]

_THIN = Side(style="thin", color="D1D5DB")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def export_to_excel(topics: List[TopicInfo], file_path: str):
    os.makedirs(os.path.dirname(os.path.abspath(file_path)), exist_ok=True)

    wb = Workbook()

    # ── Sheet 1: 상세 보고서 ───────────────────────────────────────────────
    ws = wb.active
    ws.title = "ICD_Validation_Report"
    ws.freeze_panes = "A2"          # 헤더 고정

    headers = [col[0] for col in _COLUMNS]

    # 헤더 행
    for c_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c_idx, value=h)
        cell.fill   = _FILL["header"]
        cell.font   = _FONT["header"]
        cell.border = _BORDER
        cell.alignment = _CENTER

    # 데이터 행
    for r_idx, t in enumerate(topics, 2):
        for c_idx, (_, fn, is_badge) in enumerate(_COLUMNS, 1):
            value = fn(t)
            cell  = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border    = _BORDER
            cell.alignment = _CENTER if is_badge else _LEFT

            if is_badge:
                fill, font = _badge_style(value)
                cell.fill = fill
                cell.font = font

    # 열 너비 자동 조정
    _col_widths = [30, 25, 20, 20, 14, 14, 12, 12, 12, 10, 10, 10, 10, 10, 10]
    for i, w in enumerate(_col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 20

    # ── Sheet 2: 요약 ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")

    summary_rows = [
        ("항목",        "건수"),
        ("전체 토픽 수", len(topics)),
        ("정상(Pass)",   sum(1 for t in topics if t.status == ValidationStatus.NORMAL)),
        ("QoS 불일치",   sum(1 for t in topics if t.status == ValidationStatus.QOS_MISMATCH)),
        ("주기 미달",    sum(1 for t in topics if t.status == ValidationStatus.HZ_MISMATCH)),
        ("미수신",       sum(1 for t in topics if t.status == ValidationStatus.NOT_RECEIVED)),
        ("대기중",       sum(1 for t in topics if t.status == ValidationStatus.PENDING)),
    ]

    for r_idx, (label, val) in enumerate(summary_rows, 1):
        for c_idx, v in enumerate([label, val], 1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=v)
            cell.border    = _BORDER
            cell.alignment = _CENTER
            if r_idx == 1:
                cell.fill = _FILL["header"]
                cell.font = _FONT["header"]

    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 10

    wb.save(file_path)
    return file_path
